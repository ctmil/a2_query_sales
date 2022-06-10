from odoo import tools, models, fields, api, _
from odoo.exceptions import ValidationError
import openpyxl
import base64
from datetime import date,datetime
import csv
import json
from odoo.tools.safe_eval import safe_eval


class A2AccountPadron(models.Model):
    _name = 'a2.account.padron'
    _description = 'a2.account.padron'

    # P;25032022;01042022;30042022;20000618765;D;S;N;0,00;01
    column1 = fields.Char('Column 1')
    date1 = fields.Date('date1')
    date2 = fields.Date('date2')
    date3 = fields.Date('date3')
    cuit = fields.Char('Cuit')
    column2 = fields.Char('Column 2')
    column3 = fields.Char('Column 3')
    column4 = fields.Char('Column 4')
    amount = fields.Float('Amount')
    column5 = fields.Char('Column 5')

    @api.model
    def load_a2_account_padron(self):
        csv_file = open('/tmp/PadronRGSPer042022.TXT','rt')
        csv_reader = csv.reader(csv_file, delimiter=';')
        line_count = 0
        fields = ['id','column1','date1','date2','date3','cuit','column2','column3','column4','amount','column5']
        data_lines = []
        print(str(datetime.now()))
        for row in csv_reader:
            #print(line_count, row)
            line_count = line_count + 1
            if line_count < 3000000:
                continue
            date1 = date(year=int(row[1][4:8]),month=int(row[1][2:4]),day=int(row[1][:2]))
            date2 = date(year=int(row[2][4:8]),month=int(row[2][2:4]),day=int(row[2][:2]))
            date3 = date(year=int(row[3][4:8]),month=int(row[3][2:4]),day=int(row[3][:2]))
            data_line = ['account.padron.' + str(line_count), row[0],str(date1),str(date2),str(date3),row[4],row[5],row[6],row[7],float(row[8].replace(',','.')),row[9]]
            #print(data_line)
            data_lines.append(data_line)
            if line_count % 1000 == 0:
                res = self.env['a2.account.padron'].load(fields,data_lines)
                data_lines = []
                print(line_count)
        if data_lines:
            res = self.env['a2.account.padron'].load(fields,data_lines)
        print(str(datetime.now()))
        csv_file.close()

class A2Sales(models.Model):
    _name = "a2.sales"
    _description = "A2 Sales"
    _auto = False

    partner_id = fields.Many2one('res.partner','Cliente')
    city = fields.Char('Ciudad')
    state_id = fields.Many2one('res.country.state','Provincia')
    product_id = fields.Many2one('product.product','Producto')
    product_uom_qty = fields.Float('Cantidad')

    def init(self):
        tools.drop_view_if_exists(self._cr, self._table)
        query = """
            select sol.id,so.partner_id,pa.city,pa.state_id,sol.product_id,sol.product_uom_qty
            from sale_order_line sol
            inner join sale_order so on so.id = sol.order_id
            inner join res_partner pa on pa.id = so.partner_id
            """
        self._cr.execute("""CREATE or REPLACE VIEW %s as (%s)""" % (self._table, query))

    @api.model
    def load_data(self):
        fields = ['id','name','ref','city','country_id/id','category_id/id']
        data_lines = []
        workbook = openpyxl.load_workbook("/tmp/partners.xlsx")
        # Define variable para la planilla activa
        worksheet = workbook.active
        # Itera las filas para leer los contenidos de cada celda
        rows = worksheet.rows

        for x,row in enumerate(rows):
            # Saltea la primer fila porque tiene el nombre de las columnas
            if x == 0:
                continue
            # Lee cada una de las celdas en la fila
            vals = {}
            data = []
            for i,cell in enumerate(row):
                # saltea registros con valores many2one vacios
                if cell.value == None:
                    continue
                data.append(cell.value)
            data_lines.append(data)
        print(str(datetime.now()))
        res = self.env['res.partner'].load(fields,data_lines)
        print(str(datetime.now()))
        import pdb;pdb.set_trace()

    @api.model
    def load_purchase_order(self):
        fields = ['id','partner_id/id','date_order','user_id/id','partner_ref','payment_term_id/id']
        data_lines = []
        workbook = openpyxl.load_workbook("/tmp/purchase_orders.xlsx")
        # Define variable para la planilla activa
        worksheet = workbook.active
        # Itera las filas para leer los contenidos de cada celda
        rows = worksheet.rows

        for x,row in enumerate(rows):
            # Saltea la primer fila porque tiene el nombre de las columnas
            if x == 0:
                continue
            # Lee cada una de las celdas en la fila
            vals = {}
            data = []
            for i,cell in enumerate(row):
                # saltea registros con valores many2one vacios
                if cell.value == None:
                    continue
                data.append(cell.value)
            data_lines.append(data)
        print(str(datetime.now()))
        res = self.env['purchase.order'].with_context({'tracking_disable': True}).load(fields,data_lines)
        print(str(datetime.now()))
        import pdb;pdb.set_trace()

    @api.model
    def load_purchase_order_line(self):
        fields = ['id','order_id/id','product_id/id','name','product_qty','price_unit','taxes_id/id']
        data_lines = []
        workbook = openpyxl.load_workbook("/tmp/purchase_order_lines.xlsx")
        # Define variable para la planilla activa
        worksheet = workbook.active
        # Itera las filas para leer los contenidos de cada celda
        rows = worksheet.rows

        for x,row in enumerate(rows):
            # Saltea la primer fila porque tiene el nombre de las columnas
            if x == 0:
                continue
            # Lee cada una de las celdas en la fila
            vals = {}
            data = []
            for i,cell in enumerate(row):
                # saltea registros con valores many2one vacios
                if cell.value == None:
                    continue
                data.append(cell.value)
            data_lines.append(data)
        print(str(datetime.now()))
        res = self.env['purchase.order.line'].with_context({'tracking_disable': True}).load(fields,data_lines)
        print(str(datetime.now()))
        import pdb;pdb.set_trace()

    @api.model
    def load_attachments(self):
        fields = ['id','res_model','res_id','type','mimetype','datas','name','res_name']
        orders = self.env['purchase.order'].search([])
        data_lines = []
        test_file = open('/tmp/stephen_king.png','rb')
        data_file = base64.b64encode(test_file.read())
        print(str(datetime.now()))
        for order in orders:
            data = [
                    'a2_query_sales.attachment_purchase_order_' + str(order.id),
                    'purchase.order',
                    order.id,
                    'binary',
                    'application/octet-stream',
                    data_file,
                    order.name,
                    order.name,
                    ]
            data_lines.append(data)
        res = self.env['ir.attachment'].with_context({'tracking_disable': True}).load(fields,data_lines)
        print(str(datetime.now()))


    @api.model
    def load_attachment_partners(self):
        fields = ['id','res_model','res_id','type','mimetype','datas','name','res_name']
        partners = self.env['res.partner'].search([('ref','ilike','REF')])
        data_lines = []
        test_file = open('/tmp/stephen_king.png','rb')
        data_file = base64.b64encode(test_file.read())
        print(str(datetime.now()))
        for partner in partners:
            data = [
                    'a2_query_sales.attachment_partner_' + str(partner.id),
                    'res.partner',
                    partner.id,
                    'binary',
                    'application/octet-stream',
                    data_file,
                    partner.name,
                    partner.name,
                    ]
            data_lines.append(data)
        res = self.env['ir.attachment'].with_context({'tracking_disable': True}).load(fields,data_lines)
        print(str(datetime.now()))
        import pdb;pdb.set_trace()

    @api.model
    def load_vendor_bills(self):
        fields = ['id','partner_id/id','invoice_date','user_id/id','ref','invoice_payment_term_id/id','journal_id/id','move_type']
        data_lines = []
        workbook = openpyxl.load_workbook("/tmp/vendor_bills.xlsx")
        # Define variable para la planilla activa
        worksheet = workbook.active
        # Itera las filas para leer los contenidos de cada celda
        rows = worksheet.rows

        for x,row in enumerate(rows):
            # Saltea la primer fila porque tiene el nombre de las columnas
            if x == 0:
                continue
            # Lee cada una de las celdas en la fila
            vals = {}
            data = []
            for i,cell in enumerate(row):
                data.append(cell.value)
            data_lines.append(data)
        print(str(datetime.now()))
        res = self.env['account.move'].with_context({'tracking_disable': True}).load(fields,data_lines)
        print(str(datetime.now()))
        import pdb;pdb.set_trace()

    @api.model
    def load_vendor_bill_lines(self):
        fields = ['id','move_id/id','exclude_from_invoice_tab','product_id/id','name','account_id/id','quantity','price_unit','debit','credit','tax_ids/id']
        data_lines = []
        workbook = openpyxl.load_workbook("/tmp/vendor_bill_line.xlsx")
        # Define variable para la planilla activa
        worksheet = workbook.active
        # Itera las filas para leer los contenidos de cada celda
        rows = worksheet.rows

        for x,row in enumerate(rows):
            # Saltea la primer fila porque tiene el nombre de las columnas
            if x == 0:
                continue
            if x > 3:
                continue
            # Lee cada una de las celdas en la fila
            vals = {}
            data = []
            for i,cell in enumerate(row):
                if cell.value:
                    data.append(str(cell.value))
                else:
                    data.append('')
            data_lines.append(data)
        print(str(datetime.now()))
        res = self.env['account.move.line'].with_context({'tracking_disable': True,'check_move_validity': False}).load(fields,data_lines)
        print(str(datetime.now()))
        import pdb;pdb.set_trace()

